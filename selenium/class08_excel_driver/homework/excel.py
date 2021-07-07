import openpyxl
from 配置.excel_webkey import WebKeys

excel=openpyxl.load_workbook('./data.xlsx')

sheet=excel['Sheet1']
for values in sheet.values:
    # print(values)
    if type(values[0]) is int:

        data = {}
        data['name'] = values[2]
        data['value'] = values[3]
        data['txt'] = values[4]

        for key in list(data.keys()):
            # print(key)
            if data[key] is None:
                del data[key]
        print(data)
        if values[1] == 'open_browser':
            wk = WebKeys(values[4])
        elif 'assert_text' == values[1]:
            status = getattr(wk, values[1])(**data)
            # 基于status来写入测试的结果
            if status:
                # 执行Pass写入
                sheet.cell(row=values[0] + 2, column=7).value = 'Pass'
            else:
                # 执行Failed写入
                sheet.cell(row=values[0] + 2, column=7).value = 'Failed'


        else:
            getattr(wk,values[1])(**data)
excel.save('./data.xlsx')