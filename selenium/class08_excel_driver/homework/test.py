import openpyxl

excel=openpyxl.load_workbook('./data.xlsx')
sheet=excel['Sheet1']

for values in sheet.values:
    print(values[2])

# sheet.cell(row=4,column=1).value='ass'
# excel.save('./data1.xlsx')
# print(sheet.cell(row=4,column=1).value)
# for row in sheet.values:
#     print(row)
print(excel.sheetnames)