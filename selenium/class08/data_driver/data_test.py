'''
    基于excel数据驱动来实现自动化测试
    既然关键字已封装，我可否通过对数据的读取，让它自行执行对应的关键对象？
'''
import openpyxl
from logs.log import Logger
# 访问excel内容
from openpyxl.styles import PatternFill, Font

from class07.key_word_web.keyword_web import WebKeys

# 日志
log = Logger().get_logger()
excel = openpyxl.load_workbook('../data/xxx.xlsx')
sheet = excel['Sheet3']
log.info('获取{0}内容成功，现在开始执行自动化测试......'.format(sheet))
# 读取excel内容，实现文件驱动自动化执行
for value in sheet.values:
    # 定义一个字典参数，用于接收excel中的所有参数内容
    args = {}
    # 定位方法
    args['name'] = value[2]
    # 定位路径
    args['value'] = value[3]
    # 输入内容
    args['txt'] = value[4]
    # 预期结果
    args['expect'] = value[6]
    # 基于A列进行判断是否为测试用例
    if type(value[0]) is int:
        '''
            在读取关键字时，分为几类情况：
                1. 关键字驱动类的实例化
                2. 断言类型的关键字
                    因为断言的关键字除去获取元素之外，还需要做对比，还需要写入断言结果
        '''
        # 判断是否实例化
        if value[1] == 'open_browser':
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            wk = WebKeys(value[4])
        # 断言关键字执行：首先执行断言，再判断断言是否成功
        elif 'assert' in value[1]:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            status = getattr(wk, value[1])(**args)
            # 依据status来判定写入的内容是Failed还是Pass
            if status:
                # 写入当前行的实际结果值
                # sheet.cell(row='编号+1',column=8)
                sheet.cell(row=value[0] + 1, column=8).value = 'Pass'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='AACF91')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            else:
                sheet.cell(row=value[0] + 1, column=8).value = 'Failed'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            excel.save('../data/xxx.xlsx')
        # 非特殊关键字，通过反射机制实现
        else:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            getattr(wk, value[1])(**args)
            # getattr(wk, 'open')(**args)
            # wk.open(**args)
            # 原有的open函数是 def open(self,url):
            '''
                如果是open，只需要传入value[4]
                如果是input，则需要传入value[2],value[3],value[4]
                如果是click，则需要传入value[2],value[3]
                **args不定长传参，不定义参数数量的多少，说白了就是一个字典格式
                但是字典是键值对形式存在
            '''
excel.close()
