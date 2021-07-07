'''
    excel的数据驱动实现：
        1. excel操作流程：先操作workbook，再操作sheet，再操作cell（value）
        2. 对于模块的api建议在使用前可以百度了解一下
        3. excel操作结束后记得close
'''
# 导入模块
import openpyxl

# excel操作流程
# 操作工作簿：指定文件路径，进行文件读取，类似于open()函数的文件读取操作
excel = openpyxl.load_workbook('../data/xxx.xlsx')
print(excel)
# 获取sheet：基于工作簿来获取sheet页
names = excel.sheetnames
print(names)
print(type(names))
for name in names:
    print(name)
# 操作单元格cell：指定sheet页，再进行操作
sheet = excel['Sheet3']
print(sheet)
# 获取单元格内容
# print(sheet.values)
# for value in sheet.values:
#     print(value[0])
# 指定单元格进行内容获取
# value = sheet['A5'].value
# print(value)
# 获取行：最大行数
# rows = sheet.max_row
# print(rows)
# # 获取列：最大列数
# column = sheet.max_column
# print(column)
# # 写入前一定确保被操作文件处于关闭状态，否则会报错
# # 写入：基于单元格来进行写入
# sheet['H1'] = 'asd'
# # 写入之后一定记得保存：保存对象是excel，而不是sheet或者cell
# excel.save('../data/xxx.xlsx')
# # 在所有操作结束之后，记得释放
# excel.close()
# 单元格的写入
