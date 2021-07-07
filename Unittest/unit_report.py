import  unittest
from time import sleep

import openpyxl
from 配置.excel_webkey import WebKeys
from 配置.webkey import WebKeys1
import yaml
from ddt import ddt,data,file_data,unpack
@ddt
class Unit_report(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        print('class.setUp')
    @classmethod
    def tearDownClass(cls) -> None:
        print('class.teardown')
    def setUp(self) -> None:
        print('setup')
    def tearDown(self) -> None:
        print('teardown')
    def test01(self):
        excel=openpyxl.load_workbook('./data/data.xlsx')
        sheet=excel['Sheet1']
        for values in sheet.values:
            if type(values[0]) is int:
                data={}
                data['name'] = values[2]
                data['value'] = values[3]
                data['txt'] = values[4]
                for  key in list(data.keys()):
                    if data[key] is None:
                        del data[key]
                print(values)
                if values[1]=='open_browser':
                    wk=WebKeys('Chrome')
                elif values[1]=='assert_test':
                    status = getattr(wk, values[1])(**data)
                    if status:
                        # 执行Pass写入
                        sheet.cell(row=values[0] + 2, column=7).value = 'Pass'
                    else:
                        # 执行Failed写入
                        sheet.cell(row=values[0] + 2, column=7).value = 'Failed'
                else:
                    getattr(wk,values[1])(**data)

    def test02(self):
        wk = WebKeys1('Chrome')
        wk.open(url='http://39.98.138.157/shopxo/')
        wk.click(name='link text', value='登录')
        wk.input(name='name', value='accounts', txt='xuzhu666')
        wk.input(name='name', value='pwd', txt='123456')
        wk.click(name='xpath', value='/html/body/div[4]/div/div[2]/div[2]/form/div[3]/button')
        sleep(3)
        wk.quit()
    @file_data('../Unittest/data/test_data.yaml')
    def test03(self,**kwargs):
        self.wk = WebKeys1('Chrome')
        self.wk.open(kwargs['url'])
        self.wk.input(**kwargs['input'])

    def test04(self,):
        pass
    def test05(self):
        print('05')
    def test06(self):
        print('06')

if __name__ == '__main__':
    unittest.main()
